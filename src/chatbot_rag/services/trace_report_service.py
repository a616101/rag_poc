"""
Langfuse Trace 報表生成服務。

提供 Trace 資料查詢與 Excel 報表生成功能。
"""

import asyncio
import json
import uuid
from datetime import datetime
from io import BytesIO
from typing import Optional, Union

from langfuse import get_client
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from chatbot_rag.models.report import TaskInfo, TaskStatus, TraceReportRequest


class TraceReportService:
    """Langfuse Trace 報表生成服務。"""

    SYNC_THRESHOLD = 200  # 同步處理的閾值

    def __init__(self):
        self._task_store: dict[str, TaskInfo] = {}

    def _get_langfuse(self):
        """取得 Langfuse client。"""
        return get_client()

    async def generate_report(
        self,
        request: TraceReportRequest,
    ) -> Union[bytes, TaskInfo]:
        """
        生成報表。

        - 若 trace 數量 ≤ 閾值：同步返回 Excel bytes
        - 若 trace 數量 > 閾值：建立背景任務，返回 TaskInfo

        Args:
            request: 報表請求參數

        Returns:
            bytes: Excel 檔案內容（同步模式）
            TaskInfo: 任務資訊（背景模式）
        """
        # 預先查詢 trace 數量
        count = await self._count_traces(request)
        logger.info(f"[REPORT] Estimated trace count: {count}")

        if count <= self.SYNC_THRESHOLD:
            return await self._generate_sync(request)
        else:
            return await self._create_background_task(request, count)

    async def _count_traces(self, request: TraceReportRequest) -> int:
        """預估 trace 數量。"""
        langfuse = self._get_langfuse()

        try:
            # 使用 limit=1 快速取得 meta 資訊
            result = langfuse.api.trace.list(
                limit=1,
                from_timestamp=request.start_date,
                to_timestamp=request.end_date,
                user_id=request.user_id,
                session_id=request.session_id,
                name=request.trace_name,
                tags=request.tags,
            )
            # Langfuse API 可能不提供 total count，使用 limit 作為估算
            # 實際實作中可能需要分頁計數
            return min(request.limit, 1000)  # 保守估計
        except Exception as e:
            logger.warning(f"[REPORT] Failed to count traces: {e}")
            return request.limit

    async def _generate_sync(self, request: TraceReportRequest) -> bytes:
        """同步生成 Excel 報表。"""
        logger.info("[REPORT] Starting sync report generation")

        # 1. 獲取 traces（已包含 scores）
        traces = await self._fetch_traces(request, request.score_names)
        logger.info(f"[REPORT] Fetched {len(traces)} traces")

        # 統計 scores 數量
        total_scores = sum(len(t.get("scores", [])) for t in traces)
        logger.info(f"[REPORT] Total scores in traces: {total_scores}")

        # 2. 獲取每個 trace 的 observations
        observations = await self._fetch_observations(traces)
        logger.info(f"[REPORT] Fetched {len(observations)} observations")

        # 3. 生成 Excel（scores 已整合在 traces 中）
        excel_bytes = self._create_excel(traces, observations)
        logger.info("[REPORT] Excel report generated successfully")

        return excel_bytes

    async def _create_background_task(
        self,
        request: TraceReportRequest,
        estimated_count: int,
    ) -> TaskInfo:
        """建立背景任務。"""
        task_id = uuid.uuid4().hex
        filename = f"trace_report_{request.start_date.date()}_{request.end_date.date()}.xlsx"

        task_info = TaskInfo(
            task_id=task_id,
            status=TaskStatus.PENDING,
            estimated_traces=estimated_count,
            created_at=datetime.now(),
            filename=filename,
        )
        self._task_store[task_id] = task_info

        logger.info(f"[REPORT] Created background task: {task_id}")

        # 啟動背景任務
        asyncio.create_task(self._run_background_task(task_id, request))

        return task_info

    async def _run_background_task(
        self,
        task_id: str,
        request: TraceReportRequest,
    ):
        """執行背景任務。"""
        task_info = self._task_store.get(task_id)
        if not task_info:
            return

        try:
            task_info.status = TaskStatus.PROCESSING
            logger.info(f"[REPORT] Background task {task_id} started")

            # 生成報表
            excel_bytes = await self._generate_sync(request)

            # 更新任務狀態
            task_info.status = TaskStatus.COMPLETED
            task_info.completed_at = datetime.now()
            task_info.result = excel_bytes

            logger.info(f"[REPORT] Background task {task_id} completed")

        except Exception as e:
            task_info.status = TaskStatus.FAILED
            task_info.error = str(e)
            logger.error(f"[REPORT] Background task {task_id} failed: {e}")

    def get_task(self, task_id: str) -> Optional[TaskInfo]:
        """取得任務資訊。"""
        return self._task_store.get(task_id)

    async def _fetch_traces(
        self,
        request: TraceReportRequest,
        score_names: Optional[list[str]] = None,
    ) -> list[dict]:
        """獲取 traces，包含 scores 資訊。"""
        langfuse = self._get_langfuse()
        traces = []
        page = 1
        limit = 100

        while len(traces) < request.limit:
            try:
                result = langfuse.api.trace.list(
                    limit=limit,
                    page=page,
                    from_timestamp=request.start_date,
                    to_timestamp=request.end_date,
                    user_id=request.user_id,
                    session_id=request.session_id,
                    name=request.trace_name,
                    tags=request.tags,
                )

                if not result.data:
                    break

                for trace in result.data:
                    # 獲取 trace 詳情以取得 scores
                    trace_dict = self._trace_to_dict(trace)

                    # 嘗試獲取該 trace 的 scores
                    trace_scores = self._fetch_trace_scores(
                        langfuse, trace.id, score_names
                    )
                    trace_dict["scores"] = trace_scores

                    traces.append(trace_dict)
                    if len(traces) >= request.limit:
                        break

                if len(result.data) < limit:
                    break

                page += 1

            except Exception as e:
                logger.error(f"[REPORT] Failed to fetch traces page {page}: {e}")
                break

        return traces

    def _fetch_trace_scores(
        self,
        langfuse,
        trace_id: str,
        score_names: Optional[list[str]] = None,
    ) -> list[dict]:
        """獲取單一 trace 的 scores。"""
        scores = []

        try:
            # 嘗試獲取 trace 詳情，其中可能包含 scores
            trace_detail = langfuse.api.trace.get(trace_id)

            # 檢查是否有 scores 屬性
            trace_scores = getattr(trace_detail, "scores", None)

            if trace_scores:
                for score in trace_scores:
                    score_name = getattr(score, "name", None)
                    if score_names and score_name not in score_names:
                        continue

                    scores.append({
                        "name": score_name,
                        "value": getattr(score, "value", None),
                        "comment": getattr(score, "comment", None),
                        "data_type": getattr(score, "data_type", None),
                    })

        except Exception as e:
            logger.debug(f"[REPORT] Could not fetch scores for trace {trace_id}: {e}")

        return scores

    def _trace_to_dict(self, trace) -> dict:
        """將 trace 物件轉換為 dict。"""
        # 處理 input/output，可能是 dict、list 或 string
        input_content = getattr(trace, "input", None)
        output_content = getattr(trace, "output", None)

        input_content = self._serialize_content(input_content)
        output_content = self._serialize_content(output_content)

        # 安全地處理 usage（可能不存在於 TraceWithDetails）
        usage = getattr(trace, "usage", None)
        input_tokens = 0
        output_tokens = 0
        total_tokens = 0

        if usage:
            if hasattr(usage, "input"):
                input_tokens = usage.input or 0
            elif isinstance(usage, dict):
                input_tokens = usage.get("input", 0)

            if hasattr(usage, "output"):
                output_tokens = usage.output or 0
            elif isinstance(usage, dict):
                output_tokens = usage.get("output", 0)

            if hasattr(usage, "total"):
                total_tokens = usage.total or 0
            elif isinstance(usage, dict):
                total_tokens = usage.get("total", 0)

        # 嘗試從其他屬性獲取 token 資訊
        if total_tokens == 0:
            input_tokens = getattr(trace, "input_tokens", 0) or 0
            output_tokens = getattr(trace, "output_tokens", 0) or 0
            total_tokens = getattr(trace, "total_tokens", 0) or input_tokens + output_tokens

        return {
            "id": trace.id,
            "name": getattr(trace, "name", None),
            "timestamp": getattr(trace, "timestamp", None),
            "duration": getattr(trace, "duration", None),
            "user_id": getattr(trace, "user_id", None),
            "session_id": getattr(trace, "session_id", None),
            "input": input_content,
            "output": output_content,
            "tags": getattr(trace, "tags", []) or [],
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": total_tokens,
            "total_cost": getattr(trace, "total_cost", None),
        }

    async def _fetch_observations(self, traces: list[dict]) -> list[dict]:
        """獲取所有 traces 的 observations。"""
        langfuse = self._get_langfuse()
        observations = []

        for trace in traces:
            try:
                result = langfuse.api.observations.get_many(
                    trace_id=trace["id"],
                    limit=100,
                )

                for obs in result.data:
                    observations.append(self._observation_to_dict(obs, trace["id"]))

            except Exception as e:
                logger.warning(
                    f"[REPORT] Failed to fetch observations for trace {trace['id']}: {e}"
                )

        return observations

    def _observation_to_dict(self, obs, trace_id: str) -> dict:
        """將 observation 物件轉換為 dict。"""
        # 安全地處理 usage
        usage = getattr(obs, "usage", None)
        input_tokens = 0
        output_tokens = 0

        if usage:
            if hasattr(usage, "input"):
                input_tokens = usage.input or 0
            elif isinstance(usage, dict):
                input_tokens = usage.get("input", 0)

            if hasattr(usage, "output"):
                output_tokens = usage.output or 0
            elif isinstance(usage, dict):
                output_tokens = usage.get("output", 0)

        # 嘗試從其他屬性獲取
        if input_tokens == 0:
            input_tokens = getattr(obs, "input_tokens", 0) or 0
        if output_tokens == 0:
            output_tokens = getattr(obs, "output_tokens", 0) or 0

        return {
            "trace_id": trace_id,
            "id": getattr(obs, "id", None),
            "type": getattr(obs, "type", None),
            "name": getattr(obs, "name", None),
            "model": getattr(obs, "model", None),
            "start_time": getattr(obs, "start_time", None),
            "duration": getattr(obs, "duration", None),
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "cost": getattr(obs, "total_cost", None) or getattr(obs, "calculated_total_cost", None),
            "level": getattr(obs, "level", None),
        }

    def _serialize_content(self, content) -> Optional[str]:
        """將內容序列化為字串。"""
        if content is None:
            return None
        if isinstance(content, str):
            return content
        if isinstance(content, (dict, list)):
            return json.dumps(content, ensure_ascii=False, indent=2)
        return str(content)

    def _create_excel(
        self,
        traces: list[dict],
        observations: list[dict],
    ) -> bytes:
        """生成 Excel 報表。"""
        wb = Workbook()

        # Sheet 1: Trace Summary（包含 Scores）
        self._create_summary_sheet(wb, traces)

        # Sheet 2: Observations Detail
        self._create_observations_sheet(wb, observations)

        # Sheet 3: Scores Detail（展開所有 scores）
        self._create_scores_sheet(wb, traces)

        # 轉換為 bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        traces: list[dict],
    ):
        """建立 Trace Summary sheet。"""
        ws = wb.active
        ws.title = "Trace Summary"

        # 表頭（Score 和 Score Comment 直接在 Summary 中）
        headers = [
            "Trace ID",
            "Trace Name",
            "Timestamp",
            "Duration (ms)",
            "User ID",
            "Session ID",
            "Input Tokens",
            "Output Tokens",
            "Total Tokens",
            "Total Cost ($)",
            "Tags",
            "Score",
            "Score Comment",
            "Input",
            "Output",
        ]
        self._write_header(ws, headers)

        # 資料
        for row_idx, trace in enumerate(traces, start=2):
            # 組合該 trace 的 scores
            trace_scores = trace.get("scores", [])

            # 格式化 scores 顯示
            if trace_scores:
                scores_text = "; ".join(
                    f"{s['name']}={s['value']}" for s in trace_scores if s.get("name")
                )
                comments_text = "; ".join(
                    f"[{s['name']}] {s['comment']}"
                    for s in trace_scores
                    if s.get("comment")
                )
            else:
                scores_text = ""
                comments_text = ""

            row_data = [
                trace["id"],
                trace["name"],
                self._format_datetime(trace["timestamp"]),
                trace["duration"],
                trace["user_id"],
                trace["session_id"],
                trace["input_tokens"],
                trace["output_tokens"],
                trace["total_tokens"],
                trace["total_cost"],
                ", ".join(trace["tags"]) if trace["tags"] else "",
                scores_text,
                comments_text,
                trace["input"],
                trace["output"],
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Score Comment, Input, Output 欄位設定自動換行
                if col_idx >= 12:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 調整欄寬
        self._auto_adjust_columns(ws, min_width=12, max_width=80)

    def _create_observations_sheet(self, wb: Workbook, observations: list[dict]):
        """建立 Observations Detail sheet。"""
        ws = wb.create_sheet("Observations Detail")

        # 表頭
        headers = [
            "Trace ID",
            "Observation ID",
            "Type",
            "Name",
            "Model",
            "Start Time",
            "Duration (ms)",
            "Input Tokens",
            "Output Tokens",
            "Cost ($)",
            "Level",
        ]
        self._write_header(ws, headers)

        # 資料
        for row_idx, obs in enumerate(observations, start=2):
            row_data = [
                obs["trace_id"],
                obs["id"],
                obs["type"],
                obs["name"],
                obs["model"],
                self._format_datetime(obs["start_time"]),
                obs["duration"],
                obs["input_tokens"],
                obs["output_tokens"],
                obs["cost"],
                obs["level"],
            ]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 調整欄寬
        self._auto_adjust_columns(ws, min_width=12, max_width=50)

    def _create_scores_sheet(self, wb: Workbook, traces: list[dict]):
        """建立 Scores sheet（從 traces 中提取 scores）。"""
        ws = wb.create_sheet("Scores")

        # 表頭
        headers = [
            "Trace ID",
            "Trace Name",
            "Score Name",
            "Value",
            "Data Type",
            "Comment",
        ]
        self._write_header(ws, headers)

        # 從 traces 中提取所有 scores
        row_idx = 2
        for trace in traces:
            trace_scores = trace.get("scores", [])
            for score in trace_scores:
                row_data = [
                    trace["id"],
                    trace["name"],
                    score.get("name"),
                    score.get("value"),
                    score.get("data_type"),
                    score.get("comment"),
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Comment 欄位設定自動換行
                    if col_idx == 6:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

                row_idx += 1

        # 調整欄寬
        self._auto_adjust_columns(ws, min_width=12, max_width=60)

    def _write_header(self, ws, headers: list[str]):
        """寫入表頭並套用樣式。"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 凍結首行
        ws.freeze_panes = "A2"

    def _auto_adjust_columns(
        self,
        ws,
        min_width: int = 10,
        max_width: int = 50,
    ):
        """自動調整欄寬。"""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    # 取第一行的長度（避免過長內容影響欄寬）
                    first_line = cell_value.split("\n")[0] if cell_value else ""
                    max_length = max(max_length, len(first_line))
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width

    def _format_datetime(self, dt) -> str:
        """格式化 datetime。"""
        if dt is None:
            return ""
        if isinstance(dt, str):
            return dt
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        return str(dt)


trace_report_service = TraceReportService()
